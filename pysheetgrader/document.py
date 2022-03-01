from pysheetgrader.sheet import Sheet
from pysheetgrader.utils import get_headers

from openpyxl import load_workbook

class Document:
    """
    Represents a document in the grading process. Please call the `close()` method when it's not used anymore.

    Attributes:
        GRADING_ORDER_SHEET_NAME    Holds the name of the sheet in a key workbook that contains the order
                                    of grading other sheets.
        COLUMN_HEADER_NAMES         Holds the variables (headers) to be extracted from the order sheet 
                                    and possible names (for backwards compatibility) of those variables
    """

    # How to add a new header:
    # - add the name of the header in COLUMN_HEADER_NAMES. Since it is a new header, you can just add the same to the list
    # - follow process for the "name" variable in get_grading_sheets function

    GRADING_ORDER_SHEET_NAME = 'SheetGradingOrder'
    COLUMN_HEADER_NAMES = {
        "name": ["name", "sheet", "tab"],
        "minimum_work": ["min-work"],
        "feedback": ["feedback"]
    }

    def __init__(self, path, read_only=True):
        """
        Initializer for this class.
        :param path: Valid path of the document. This path will be opened into two workbooks: `formula_wb` and
            `computed_value_wb`.
        :param read_only: Boolean marker whether the document should be treated as read-only. This will affect
            the `formula_wb` and `computed_value_wb` property of this instance - whether they're read-only or not.
            Please set this to `False` when creating key documents, so the rubric notes can be accessed.
            Defaults to `True`.
        """

        self.path = path
        self.read_only = read_only

        self.formula_wb = load_workbook(path, read_only=read_only, data_only=False)
        self.computed_value_wb = load_workbook(path, read_only=read_only, data_only=True)

    def is_valid_key(self):
        """
        Returns a Boolean value to identify whether this document is a valid key document or not.
        :return: Boolean value.
        """
        return self.GRADING_ORDER_SHEET_NAME in self.formula_wb.sheetnames

    def get_grading_sheets(self) -> [Sheet]:
        """
        Returns ordered list of sheets to be graded (not including the GradingOrderSheet).
        Will return an empty array if the `is_valid_key` is False.

        :return: List of Sheet.
        """

        # Early return
        if not self.is_valid_key():
            return []

        sheets = []
        order_sheet = self.formula_wb[self.GRADING_ORDER_SHEET_NAME]

        first_row = next(order_sheet.iter_rows(values_only=True))
        header_index = get_headers(self.COLUMN_HEADER_NAMES, first_row)
        has_name = "name" in header_index
        has_minimum_work = "minimum_work" in header_index
        has_feedback = "feedback" in header_index

        for row in order_sheet.iter_rows(min_row=2, values_only=True):
            name = row[header_index["name"]] if has_name else None
            minimum_work = row[header_index["minimum_work"]] if has_minimum_work else None
            feedback = row[header_index["feedback"]] if has_feedback else None

            sheets.append(Sheet( name, minimum_work, feedback))

        return sheets

    def close(self):
        """
        Closes this document. Call this method only after the document is not used anymore.
        """
        self.formula_wb.close()
        self.computed_value_wb.close()

    def save(self):
        # self.formula_wb.save(self.path)
        self.computed_value_wb.save(self.path)

