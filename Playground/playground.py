from openpyxl import load_workbook
import os
import sys
import xlwings as xw

class Sheet:
    """
    Represents a sheet in the grading process.
    """

    def __init__(self, name: str, minimum_work: float, feedback: str):
        """
        Initialize a sheet object representing a graded sheet

        :param name: the name of the sheet
        :param minimum_work: an int, if the student's score below this, the grading will abort
        :param feedback: the feedback string when student doesn't achieve the minimum work
        """
        self.name = name if name is not None else ""
        self.minimum_work = minimum_work if minimum_work is not None else 0.0
        self.feedback = feedback if feedback is not None else ""

class Document:
 
    def __init__(self, path, read_only=True):
      
        self.path = path
        self.read_only = read_only

        self.formula_wb = load_workbook(path, read_only=read_only, data_only=False)
        self.computed_value_wb = load_workbook(path, read_only=read_only, data_only=True)
        # self.graded_path = self.path[:-5]+"_graded"+self.path[-5:]

key_doc = Document("key.xlsx",read_only=False)
key_sheet = key_doc.formula_wb["keysheet"]

sub_doc = Document("sub.xlsx",read_only=False)
sub_sheet = sub_doc.computed_value_wb["subsheet"]

formula = key_sheet['A3'].value
print("Formula = ", formula)
sub_sheet["R3"].value = formula
sub_doc.computed_value_wb.save("sub.xlsx")

wbxl = xw.Book('sub.xlsx')
val = wbxl.sheets['subsheet'].range('R3').value
print('Val = ',val)





